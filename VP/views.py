import csv
import json
import os
import re
from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import login, authenticate, logout as auth_logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Count, Case, When, Value, IntegerField
from django.db.models.functions import TruncMonth
from django.core.mail import send_mail
from django.conf import settings
from django.views.decorators.csrf import csrf_protect

from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, Alignment as XLAlignment, PatternFill as XLPatternFill, Border as XLBorder, Side as XLSide
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import Panel, Member, Advisor, Event, EventResult, Achievement, UserProfile, EventRegistration
from .forms import RegistrationForm, UserUpdateForm, UserProfileForm, LoginForm

# --- Helper Functions ---

ROLE_ORDER_MAP = {
    "President": 0,
    "Vice President": 1,
    "General Secretary": 2,
    "Joint Secretary": 3,
    "Treasurer": 4,
    "Assistant Treasurer": 5,
    "Organizing Secretary": 6,
    "Assistant Organizing Secretary": 7,
    "Media & Publication Secretary": 8,
    "Assistant Media & Publication Secretary": 9,
    "Executive Member": 10,
    "Member": 11,
    "General Member": 11,
}


def panel_year_sort_value(panel_year):
    """Return a numeric sort key based on the leading year in a panel label."""
    match = re.search(r"\d{4}", str(panel_year or ""))
    return int(match.group()) if match else 0


def ordered_members_for_export(queryset):
    """Serialize members by their role order first, then by panel year descending, then by name."""
    members = list(queryset.select_related("panel"))
    members.sort(
        key=lambda member: (
            ROLE_ORDER_MAP.get(member.role, 99),
            -panel_year_sort_value(member.panel.year if member.panel else ""),
            member.panel.year if member.panel else "",
            (member.name or "").lower(),
            member.pk or 0,
        )
    )
    return members


def is_admin(user):
    """Helper to check if the user is authorized as admin."""
    if not user.is_authenticated:
        return False

    # Django superusers/staff should always pass admin checks
    if user.is_superuser or user.is_staff:
        return True

    try:
        return user.userprofile.user_type == "admin"
    except UserProfile.DoesNotExist:
        return False


# --- Public Views ---


def index(request):
    """
    Landing page with panel info and a full list of live/upcoming operations.
    Updated to support horizontal scrolling for all active events.
    """
    # 1. Fetch all panels ordered by year for the recruitment/history section
    panels = Panel.objects.all().order_by("-year")

    # 2. Fetch all events that are not yet finished (Upcoming or Ongoing)
    # The slice [:3] is removed to allow all active cards to show in the scrollable row
    upcoming_events = Event.objects.filter(status__in=["Upcoming", "Ongoing"]).order_by(
        "date"
    )

    # 3. Keep the latest completed events in context for potential "Past Missions" logs
    completed_events = Event.objects.filter(status="Completed").order_by("-date")[:3]

    context = {
        "panels": panels,
        "upcoming_events": upcoming_events,
        "completed_events": completed_events,
    }
    return render(request, "index.html", context)


def panels_view(request):
    """View for displaying all panels."""
    return render(request, "VP/panels.html", {"panels": Panel.objects.all()})


def panel_detail(request, panel_id):
    """Detailed view for a specific panel roster."""
    panel = get_object_or_404(Panel, id=panel_id)
    role_order_cases = [When(role=role, then=Value(index)) for role, index in ROLE_ORDER_MAP.items()]
    members = panel.members.all().annotate(
        role_rank=Case(*role_order_cases, default=Value(99), output_field=IntegerField())
    ).order_by("role_rank", "name")
    return render(
        request,
        "VP/panel_detail.html",
        {"panel": panel, "members": members},
    )


def events_view(request):
    """Refined search and filtering logic."""
    status_filter = request.GET.get("status", "all")
    search_query = request.GET.get("search", "").strip()

    # Base queryset for search and counts
    base_events = Event.objects.all()

    # Refined Search: Title, Location, and Description
    if search_query:
        base_events = base_events.filter(
            Q(title__icontains=search_query)
            | Q(location__icontains=search_query)
            | Q(description__icontains=search_query)
        )

    # Counts for filter tabs (respect search query)
    all_count = base_events.count()
    upcoming_count = base_events.filter(status="Upcoming").count()
    ongoing_count = base_events.filter(status="Ongoing").count()
    completed_count = base_events.filter(status="Completed").count()

    # Apply status filter to final listing
    events = base_events.order_by("-date")
    if status_filter != "all":
        events = events.filter(status=status_filter)

    # Countdown section data
    upcoming_events = Event.objects.filter(status__in=["Upcoming", "Ongoing"]).order_by(
        "date"
    )

    context = {
        "events": events,
        "search_query": search_query,
        "status_filter": status_filter,
        "all_count": all_count,
        "upcoming_count": upcoming_count,
        "ongoing_count": ongoing_count,
        "completed_count": completed_count,
        "upcoming_events": upcoming_events,
    }
    return render(request, "VP/events.html", context)


def event_detail(request, event_id):
    """Detailed view for a specific event with similar events logic."""
    event = get_object_or_404(Event, id=event_id)

    # Logic to fetch up to 3 similar events based on status
    # excluding the current event itself
    similar_events = (
        Event.objects.filter(status=event.status)
        .exclude(id=event.id)
        .order_by("-date")[:3]
    )

    is_registered = False
    registration = None
    if request.user.is_authenticated:
        registration = EventRegistration.objects.filter(user=request.user, event=event).first()
        if registration:
            if registration.status == 'Rejected':
                is_registered = False
            else:
                is_registered = True
                # Retrofit QR code on load if it's missing (e.g. for existing registrations)
                if registration.status == 'Approved' and not registration.qr_code:
                    registration.generate_qr_code()
                    registration.refresh_from_db()

    context = {
        "event": event,
        "similar_events": similar_events,
        "is_registered": is_registered,
        "registration": registration,
        "is_registration_open": event.is_registration_open,
    }
    return render(request, "VP/event_detail.html", context)


def event_photos(request, event_id):
    """Gallery view for a specific mission/event."""
    event = get_object_or_404(Event, id=event_id)
    photos = event.photos.all()
    return render(
        request,
        "VP/event_photos.html",
        {"event": event, "photos": photos},
    )


def event_results(request, event_id):
    """Winner standings for a specific mission/event."""
    event = get_object_or_404(Event, id=event_id)
    results = event.results.all().order_by("order", "id")

    ranked_results = []
    for rank_value, rank_label in EventResult.RANK_CHOICES:
        rank_items = [result for result in results if result.rank == rank_value]
        if rank_items:
            ranked_results.append(
                {
                    "rank": rank_label,
                    "results": rank_items,
                }
            )

    return render(
        request,
        "VP/event_results.html",
        {
            "event": event,
            "ranked_results": ranked_results,
        },
    )


def advisors_view(request):
    """View for faculty advisors."""
    return render(request, "VP/advisors.html", {"advisors": Advisor.objects.all()})


def advisor_detail(request, advisor_id):
    """Detailed view for a specific faculty advisor."""
    advisor = get_object_or_404(Advisor, id=advisor_id)
    return render(request, "VP/advisor_detail.html", {"advisor": advisor})


def members_view(request):
    """Filterable directory of all society members."""
    role_order_cases = [When(role=role, then=Value(index)) for role, index in ROLE_ORDER_MAP.items()]
    members = Member.objects.all().select_related("panel").annotate(
        role_rank=Case(*role_order_cases, default=Value(99), output_field=IntegerField())
    ).order_by("role_rank", "-panel__year", "name")
    selected_panel = request.GET.get("panel", "all")
    selected_role = request.GET.get("role", "all")

    if selected_panel != "all":
        members = members.filter(panel_id=selected_panel)
    if selected_role != "all":
        members = members.filter(role=selected_role)

    president_count = members.filter(role="President").count()

    context = {
        "members": members,
        "panels": Panel.objects.all(),
        "selected_panel": selected_panel,
        "selected_role": selected_role,
        "president_count": president_count,
    }
    return render(request, "VP/members.html", context)


# --- Authentication Views ---


def register_view(request):
    """Handles new user registration and profile initialization."""
    if request.user.is_authenticated:
        return redirect("admin_dashboard" if is_admin(request.user) else "index")

    if request.method == "POST":
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            is_bars_member = form.cleaned_data.get("is_bars_member", "yes") == "yes"
            position_name = form.cleaned_data.get("position_name")
            UserProfile.objects.create(
                user=user,
                user_type=(
                    "member"
                    if is_bars_member
                    else form.cleaned_data.get("user_type", "student")
                ),
                student_id=form.cleaned_data.get("student_id", ""),
                phone=form.cleaned_data.get("phone", ""),
                is_bars_member=is_bars_member,
                position_name=position_name if is_bars_member else None,
            )
            login(request, user)
            messages.success(request, "Registration successful! Welcome to BARS.")
            return redirect("admin_dashboard" if is_admin(user) else "user_dashboard")
    else:
        form = RegistrationForm()
    return render(request, "VP/register.html", {"form": form})


def login_view(request):
    """Handles terminal access for authorized users."""
    if request.user.is_authenticated:
        return redirect("admin_dashboard" if is_admin(request.user) else "index")

    if request.method == "POST":
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f"Welcome back, {user.username}!")
            if is_admin(user):
                return redirect("admin_dashboard")
            return redirect("user_dashboard")
    else:
        form = LoginForm()
    return render(request, "VP/login.html", {"form": form})


def logout_view(request):
    """Terminates session and clears local data."""
    auth_logout(request)
    messages.success(request, "LOGOUT SUCCESSFUL. SESSION TERMINATED.")
    return redirect("index")


@login_required
def user_dashboard(request):
    """Mission dashboard for authenticated non-admin users."""
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user, user_type="guest")

    member_record = (
        Member.objects.filter(user=request.user).select_related("panel").first()
    )
    active_panel = Panel.objects.all().first()
    active_member_count = (
        Member.objects.filter(panel=active_panel).count() if active_panel else 0
    )
    upcoming_events = Event.objects.filter(status__in=["Upcoming", "Ongoing"]).order_by(
        "date"
    )[:4]
    latest_achievements = Achievement.objects.all()[:3]

    registered_event_ids = set()
    if request.user.is_authenticated:
        registered_event_ids = set(
            request.user.event_registrations.values_list("event_id", flat=True)
        )

    context = {
        "profile": profile,
        "member_record": member_record,
        "active_panel": active_panel,
        "upcoming_events": upcoming_events,
        "latest_achievements": latest_achievements,
        "total_events": Event.objects.count(),
        "total_members": active_member_count,
        "panel_count": Panel.objects.count(),
        "registered_event_ids": registered_event_ids,
    }
    return render(request, "VP/user_dashboard.html", context)


@login_required
def user_profile(request):
    """User management console for personal data updates."""
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user, user_type="guest")

    if request.method == "POST":
        form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = UserProfileForm(request.POST, instance=profile)
        if form.is_valid() and profile_form.is_valid():
            form.save()
            profile_form.save()
            messages.success(request, "Your profile has been updated!")
            return redirect("user_profile")
    else:
        form = UserUpdateForm(instance=request.user)
        profile_form = UserProfileForm(instance=profile)

    return render(
        request, "VP/user_profile.html", {"form": form, "profile_form": profile_form}
    )


# --- Administrative Terminal Views ---


@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    """
    Administrative Command Deck.
    Includes Growth Analytics, Operational Timers, and System Logs.
    """
    panels = Panel.objects.all().order_by("-year")
    members = Member.objects.all()
    events = Event.objects.all()

    # 2. Operational Awareness
    upcoming_schedule = events.filter(status="Upcoming").order_by("date")

    context = {
        "panels": panels,
        "total_members": members.count(),
        "total_events": events.count(),
        "upcoming_ops": upcoming_schedule.count(),
        "active_panel": panels.first() if panels.exists() else None,
        "recent_registrations": members.order_by("-id")[:5],
        "upcoming_schedule": upcoming_schedule[:5],
        "event_registrations": EventRegistration.objects.all().order_by("-registered_at"),
    }
    return render(request, "VP/admin_dashboard.html", context)


@login_required
@user_passes_test(is_admin)
def export_members_csv(request):
    """Generates a CSV report of society roster for official use."""
    response = HttpResponse(content_type="text/csv")
    filename = f"BARS_Roster_{datetime.now().strftime('%Y-%m-%d')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(["FULL NAME", "DESIGNATION", "PANEL YEAR", "CONTACT EMAIL"])

    members = ordered_members_for_export(Member.objects.all())
    for m in members:
        writer.writerow([m.name, m.role, m.panel.year, m.email])

    return response


def about_view(request):
    return render(request, "VP/about.html")


def achievements_view(request):
    achievements = Achievement.objects.all()
    return render(request, "VP/achievements.html", {"achievements": achievements})


def developers_view(request):
    return render(request, "VP/developers.html")


@csrf_protect  # Ensures security is active
def submit_triumph(request):
    if request.method == "POST":
        # Extract data from the POST request
        name = request.POST.get("name", "").strip()
        title = request.POST.get("title", "").strip()
        category = request.POST.get("category", "")
        description = request.POST.get("description", "").strip()

        # Simple validation check
        if not name or not title or not description:
            return JsonResponse(
                {"status": "error", "message": "All fields required."}, status=400
            )

        subject = f"BARS MAINFRAME: New Triumph by {name}"
        message = f"Commander, a victory was reported.\n\nMember: {name}\nTitle: {title}\nCategory: {category}\n\nDescription: {description}"

        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [settings.EMAIL_HOST_USER],
                fail_silently=False,
            )
            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse(
                {"status": "error", "message": "SMTP Uplink Failed."}, status=500
            )

    return JsonResponse(
        {"status": "error", "message": "Invalid request method."}, status=405
    )


@csrf_protect
def submit_event_proposal(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        email = request.POST.get("email", "").strip()
        event_name = request.POST.get("event_name", "").strip()
        category = request.POST.get("category", "").strip()
        details = request.POST.get("details", "").strip()

        if not name or not email or not event_name or not category or not details:
            return JsonResponse(
                {"status": "error", "message": "All fields required."}, status=400
            )

        subject = f"BARS EVENT PROPOSAL: {event_name}"
        message = (
            "Commander, a new event proposal has been submitted.\n\n"
            f"Name: {name}\n"
            f"Email: {email}\n"
            f"Event Name: {event_name}\n"
            f"Category: {category}\n\n"
            "Details:\n"
            f"{details}"
        )

        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [settings.EMAIL_HOST_USER],
                fail_silently=False,
            )
            return JsonResponse({"status": "success"})
        except Exception:
            return JsonResponse(
                {"status": "error", "message": "SMTP Uplink Failed."}, status=500
            )

    return JsonResponse(
        {"status": "error", "message": "Invalid request method."}, status=405
    )


# --- Admin Intelligence Export Terminal & Document Generators ---

class NumberedCanvas(canvas.Canvas):
    """Canvas class to generate 'Page X of Y' for ReportLab PDF documents."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#666666"))
        
        # Bottom rule divider
        self.setStrokeColor(colors.HexColor("#00F3FF"))
        self.setLineWidth(0.5)
        self.line(36, 45, 576, 45) # 0.5in margins: letter width 612 -> page width 540
        
        # Footer text
        self.drawString(36, 30, "BAIUST Robotics Society - SECURE REPORT")
        self.drawRightString(576, 30, f"Page {self._pageNumber} of {page_count}")
        self.restoreState()


def generate_csv_response(resource_name, queryset):
    response = HttpResponse(content_type="text/csv")
    filename = f"BARS_{resource_name}_{datetime.now().strftime('%Y-%m-%d')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    if resource_name == "members":
        writer.writerow(["FULL NAME", "ROLE", "DEPARTMENT", "PANEL YEAR", "EMAIL", "MOBILE NUMBER", "LINKEDIN", "GITHUB"])
        for m in queryset:
            writer.writerow([m.name, m.role, m.get_department_display(), m.panel.year if m.panel else '', m.email, m.mobile_number or '', m.linkedin, m.github])
    elif resource_name == "events":
        writer.writerow(["TITLE", "DATE", "END DATE", "LOCATION", "STATUS", "DESCRIPTION"])
        for e in queryset:
            writer.writerow([e.title, e.date.strftime('%Y-%m-%d %H:%M') if e.date else '', e.end_date.strftime('%Y-%m-%d %H:%M') if e.end_date else '', e.location, e.status, e.description])
    elif resource_name == "registrations":
        writer.writerow(["USERNAME", "FULL NAME", "EMAIL", "USER TYPE", "STUDENT ID", "PHONE", "IS BARS MEMBER", "POSITION NAME", "CREATED AT"])
        for u in queryset:
            writer.writerow([u.user.username, u.user.get_full_name(), u.user.email, u.get_user_type_display(), u.student_id, u.phone, "Yes" if u.is_bars_member else "No", u.position_name or '', u.created_at.strftime('%Y-%m-%d %H:%M') if u.created_at else ''])
            
    return response


def generate_json_response(resource_name, queryset):
    data = []
    if resource_name == "members":
        for m in queryset:
            data.append({
                "name": m.name,
                "role": m.role,
                "department": m.get_department_display(),
                "panel_year": m.panel.year if m.panel else '',
                "email": m.email,
                "mobile_number": m.mobile_number or '',
                "linkedin": m.linkedin,
                "github": m.github
            })
    elif resource_name == "events":
        for e in queryset:
            data.append({
                "title": e.title,
                "date": e.date.strftime('%Y-%m-%d %H:%M') if e.date else '',
                "end_date": e.end_date.strftime('%Y-%m-%d %H:%M') if e.end_date else '',
                "location": e.location,
                "status": e.status,
                "description": e.description
            })
    elif resource_name == "registrations":
        for u in queryset:
            data.append({
                "username": u.user.username,
                "full_name": u.user.get_full_name(),
                "email": u.user.email,
                "user_type": u.get_user_type_display(),
                "student_id": u.student_id,
                "phone": u.phone,
                "is_bars_member": u.is_bars_member,
                "position_name": u.position_name or '',
                "created_at": u.created_at.strftime('%Y-%m-%d %H:%M') if u.created_at else ''
            })
            
    response = HttpResponse(json.dumps(data, indent=4), content_type="application/json")
    filename = f"BARS_{resource_name}_{datetime.now().strftime('%Y-%m-%d')}.json"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def generate_excel_response(resource_name, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = resource_name.capitalize()
    ws.views.sheetView[0].showGridLines = True
    
    font_title = XLFont(name="Segoe UI", size=14, bold=True, color="FF6B00")
    font_header = XLFont(name="Segoe UI", size=11, bold=True, color="00F3FF")
    fill_header = XLPatternFill(start_color="001F3F", end_color="001F3F", fill_type="solid")
    alignment_center = XLAlignment(horizontal="center", vertical="center")
    alignment_left = XLAlignment(horizontal="left", vertical="center")
    border_thin = XLBorder(
        left=XLSide(style='thin', color='CCCCCC'),
        right=XLSide(style='thin', color='CCCCCC'),
        top=XLSide(style='thin', color='CCCCCC'),
        bottom=XLSide(style='thin', color='CCCCCC')
    )
    
    # Title Block
    ws.append([f"BAIUST ROBOTICS SOCIETY - {resource_name.upper()} DATA EXPORT"])
    ws.row_dimensions[1].height = 35
    ws.merge_cells("A1:H1")
    ws["A1"].font = font_title
    ws["A1"].alignment = alignment_center
    ws["A1"].fill = XLPatternFill(start_color="0D0D0D", end_color="0D0D0D", fill_type="solid")
    
    ws.append([]) # Spacer row
    
    if resource_name == "members":
        headers = ["Full Name", "Role", "Department", "Panel Year", "Email", "Mobile Number", "LinkedIn", "GitHub"]
    elif resource_name == "events":
        headers = ["Title", "Date", "End Date", "Location", "Status", "Description"]
    elif resource_name == "registrations":
        headers = ["Username", "Full Name", "Email", "User Type", "Student ID", "Phone", "Is BARS Member", "Position Name", "Created At"]
    else:
        headers = []
        
    ws.append(headers)
    header_row_idx = 3
    ws.row_dimensions[header_row_idx].height = 25
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = alignment_center
        cell.border = border_thin
        
    row_idx = 4
    if resource_name == "members":
        for m in queryset:
            row_data = [m.name, m.role, m.get_department_display(), m.panel.year if m.panel else '', m.email, m.mobile_number or '', m.linkedin, m.github]
            ws.append(row_data)
            row_idx += 1
    elif resource_name == "events":
        for e in queryset:
            row_data = [e.title, e.date.strftime('%Y-%m-%d %H:%M') if e.date else '', e.end_date.strftime('%Y-%m-%d %H:%M') if e.end_date else '', e.location, e.status, e.description]
            ws.append(row_data)
            row_idx += 1
    elif resource_name == "registrations":
        for u in queryset:
            row_data = [u.user.username, u.user.get_full_name(), u.user.email, u.get_user_type_display(), u.student_id, u.phone, "Yes" if u.is_bars_member else "No", u.position_name or '', u.created_at.strftime('%Y-%m-%d %H:%M') if u.created_at else '']
            ws.append(row_data)
            row_idx += 1
            
    fill_even = XLPatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    fill_odd = XLPatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for r in range(4, row_idx):
        ws.row_dimensions[r].height = 20
        is_even = (r % 2 == 0)
        row_fill = fill_even if is_even else fill_odd
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border_thin
            cell.fill = row_fill
            cell.alignment = alignment_left
            cell.font = XLFont(name="Segoe UI", size=10)

    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"BARS_{resource_name}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def generate_pdf_response(resource_name, queryset):
    response = HttpResponse(content_type="application/pdf")
    filename = f"BARS_{resource_name}_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    style_title = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#001F3F"),
        alignment=0
    )
    style_subtitle = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#666666")
    )
    style_cell = ParagraphStyle(
        "CellText",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#333333")
    )
    style_cell_bold = ParagraphStyle(
        "CellTextBold",
        parent=style_cell,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#001F3F")
    )
    style_header = ParagraphStyle(
        "TableHeaderText",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#00F3FF"),
        alignment=1
    )
    
    logo_path = os.path.join(settings.BASE_DIR, 'media', 'logo.png')
    
    title_text = f"BAIUST ROBOTICS SOCIETY<br/><font color='#FF6B00' size='13'><b>{resource_name.upper()} OFFICIAL REPORT</b></font>"
    title_para = Paragraph(title_text, style_title)
    
    meta_text = f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}<br/><b>Status:</b> SECURE ONLINE EXTRACT | BARS MAINFRAME"
    meta_para = Paragraph(meta_text, style_subtitle)
    
    left_cell = [title_para, Spacer(1, 4), meta_para]
    
    if os.path.exists(logo_path):
        try:
            logo_img = Image(logo_path, width=45, height=45)
            header_table = Table([[left_cell, logo_img]], colWidths=[465, 75])
        except Exception:
            header_table = Table([[left_cell, '']], colWidths=[465, 75])
    else:
        header_table = Table([[left_cell, '']], colWidths=[465, 75])
        
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    
    story.append(header_table)
    
    divider = Table([['']], colWidths=[540], rowHeights=[2])
    divider.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#00F3FF")),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(divider)
    story.append(Spacer(1, 12))
    
    table_data = []
    if resource_name == "members":
        headers = ["Full Name", "Role / Designation", "Dept", "Panel", "Email", "Mobile"]
        col_widths = [110, 110, 50, 50, 130, 90]
        table_data.append([Paragraph(h, style_header) for h in headers])
        for m in queryset:
            table_data.append([
                Paragraph(m.name, style_cell_bold),
                Paragraph(m.role, style_cell),
                Paragraph(m.get_department_display(), style_cell),
                Paragraph(m.panel.year if m.panel else '', style_cell),
                Paragraph(m.email, style_cell),
                Paragraph(m.mobile_number or '', style_cell),
            ])
            
    elif resource_name == "events":
        headers = ["Title", "Date", "Location", "Status", "Description"]
        col_widths = [110, 85, 90, 60, 195]
        table_data.append([Paragraph(h, style_header) for h in headers])
        for e in queryset:
            desc = e.description
            if len(desc) > 150:
                desc = desc[:147] + "..."
            table_data.append([
                Paragraph(e.title, style_cell_bold),
                Paragraph(e.date.strftime('%Y-%m-%d %H:%M') if e.date else '', style_cell),
                Paragraph(e.location, style_cell),
                Paragraph(e.status, style_cell),
                Paragraph(desc, style_cell),
            ])
            
    elif resource_name == "registrations":
        headers = ["Username", "Full Name", "Email", "Type", "Student ID", "Phone", "Member"]
        col_widths = [75, 95, 110, 50, 55, 95, 60]
        table_data.append([Paragraph(h, style_header) for h in headers])
        for u in queryset:
            table_data.append([
                Paragraph(u.user.username, style_cell_bold),
                Paragraph(u.user.get_full_name(), style_cell),
                Paragraph(u.user.email, style_cell),
                Paragraph(u.get_user_type_display(), style_cell),
                Paragraph(u.student_id, style_cell),
                Paragraph(u.phone, style_cell),
                Paragraph("Yes" if u.is_bars_member else "No", style_cell),
            ])
            
    report_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    ts = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#001F3F")),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
    ])
    
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            ts.add('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F8FAFC"))
        else:
            ts.add('BACKGROUND', (0, i), (-1, i), colors.white)
            
    report_table.setStyle(ts)
    story.append(report_table)
    
    doc.build(story, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_admin)
def export_data(request):
    """Unified API endpoint to export various society resources in different file formats."""
    resource = request.GET.get("resource", "members")
    fmt = request.GET.get("format", "csv")
    panel_id = request.GET.get("panel")
    
    if resource == "members":
        queryset = Member.objects.all().select_related("panel")
        if panel_id:
            queryset = queryset.filter(panel_id=panel_id)
        queryset = ordered_members_for_export(queryset)
    elif resource == "events":
        queryset = Event.objects.all().order_by("-date")
    elif resource == "registrations":
        queryset = UserProfile.objects.all().select_related("user", "panel").order_by("-created_at")
        if panel_id:
            queryset = queryset.filter(panel_id=panel_id)
    else:
        return HttpResponse("Invalid resource specified", status=400)
        
    if fmt == "csv":
        return generate_csv_response(resource, queryset)
    elif fmt == "json":
        return generate_json_response(resource, queryset)
    elif fmt == "excel" or fmt == "xlsx":
        return generate_excel_response(resource, queryset)
    elif fmt == "pdf":
        return generate_pdf_response(resource, queryset)
    else:
        return HttpResponse("Invalid format specified", status=400)


def generate_ticket_pdf(registration):
    """Generates a custom PDF bytes representation of the registration ticket stub."""
    from io import BytesIO
    import os
    
    buffer = BytesIO()
    
    # Custom ticket shape page: 450x200 points
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(450, 200),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles to match the cyber website theme
    style_event = ParagraphStyle(
        "TicketEvent",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        textColor=colors.HexColor("#FFFFFF")
    )
    style_label = ParagraphStyle(
        "TicketLabel",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        textColor=colors.HexColor("#8892B0")
    )
    style_value = ParagraphStyle(
        "TicketValue",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#FFFFFF")
    )
    style_serial = ParagraphStyle(
        "TicketSerial",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#00F3FF")
    )
    
    left_flow = []
    left_flow.append(Paragraph("<b>TICKET</b>", ParagraphStyle("Header", fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#00F3FF"))))
    left_flow.append(Spacer(1, 4))
    left_flow.append(Paragraph(registration.event.title, style_event))
    left_flow.append(Spacer(1, 10))
    
    info_data = [
        [Paragraph("ATTENDEE", style_label), Paragraph("PHONE NUMBER", style_label)],
        [Paragraph(registration.name or registration.user.username, style_value), Paragraph(registration.phone or "N/A", style_value)],
        [Paragraph("SERIAL NO", style_label), Paragraph("DATE & TIME", style_label)],
        [Paragraph(f"#{registration.serial_no}", style_serial), Paragraph(registration.event.date.strftime('%Y-%m-%d %H:%M'), style_value)]
    ]
    info_table = Table(info_data, colWidths=[130, 130])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,0), (-1,-1), 2),
    ]))
    left_flow.append(info_table)
    
    qr_img = None
    if registration.qr_code and os.path.exists(registration.qr_code.path):
        try:
            # Scale down the image slightly so it fits beautifully in the right column
            qr_img = Image(registration.qr_code.path, width=100, height=100)
        except Exception:
            pass
            
    layout_data = [[left_flow, qr_img or ""]]
    layout_table = Table(layout_data, colWidths=[270, 130])
    layout_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
        # Dashed line separator between left details and QR code
        ('LINEAFTER', (0,0), (0,0), 1, colors.HexColor("#00F3FF"), 0, (3,3)),
    ]))
    
    story.append(layout_table)
    
    # Callback to draw cyber-themed dark background and dashed outer border
    def draw_cyber_background(canvas, doc):
        canvas.saveState()
        # Cyber dark background: #0A0E17
        canvas.setFillColor(colors.HexColor("#0A0E17"))
        canvas.rect(0, 0, 450, 200, fill=1, stroke=0)
        
        # Cyber cyan dashed border
        canvas.setStrokeColor(colors.HexColor("#00F3FF"))
        canvas.setLineWidth(1)
        canvas.setDash([4, 4])
        canvas.rect(10, 10, 430, 180, fill=0, stroke=1)
        
        # Tiny corner accents
        canvas.setLineWidth(1.5)
        canvas.setDash([]) # Reset to solid line
        # Top-left accent
        canvas.line(8, 192, 18, 192)
        canvas.line(8, 192, 8, 182)
        # Top-right accent
        canvas.line(442, 192, 432, 192)
        canvas.line(442, 192, 442, 182)
        # Bottom-left accent
        canvas.line(8, 8, 18, 8)
        canvas.line(8, 8, 8, 18)
        # Bottom-right accent
        canvas.line(442, 8, 432, 8)
        canvas.line(442, 8, 442, 18)
        
        canvas.restoreState()
        
    doc.build(story, onFirstPage=draw_cyber_background)
    
    buffer.seek(0)
    return buffer.getvalue()


@login_required
@csrf_protect
def register_event(request, event_id):
    """Handles event registration for authenticated users."""
    if request.method != "POST":
        return redirect("event_detail", event_id=event_id)
        
    event = get_object_or_404(Event, id=event_id)
    
    # Check if registration is open
    if not event.is_registration_open:
        if event.capacity is not None and event.registration_count >= event.capacity:
            messages.error(request, "Registration failed: This event is already full.")
        else:
            messages.error(request, "Registration is closed for this event.")
        return redirect("event_detail", event_id=event_id)
        
    # Check if already registered
    existing_registration = EventRegistration.objects.filter(user=request.user, event=event).first()
    if existing_registration:
        if existing_registration.status == 'Pending':
            messages.warning(request, "Your registration is pending approval.")
            return redirect("event_detail", event_id=event_id)
        elif existing_registration.status == 'Approved':
            messages.warning(request, "You are already registered for this event.")
            return redirect("event_detail", event_id=event_id)
        
    # Register the user
    name = request.POST.get("name", "").strip()
    student_id = request.POST.get("student_id", "").strip()
    email = request.POST.get("email", "").strip()
    phone = request.POST.get("phone", "").strip()
    payment_method = request.POST.get("payment_method", "hand_cash").strip()
    transaction_id = request.POST.get("transaction_id", "").strip()
    hand_cash_recipient = request.POST.get("hand_cash_recipient", "").strip()
    photo = request.FILES.get("photo")

    if existing_registration and existing_registration.status == 'Rejected':
        # Re-submit: update the existing record
        existing_registration.name = name
        existing_registration.student_id = student_id
        existing_registration.email = email
        existing_registration.phone = phone
        existing_registration.payment_method = payment_method
        existing_registration.transaction_id = transaction_id if payment_method == "bkash" else None
        existing_registration.hand_cash_recipient = hand_cash_recipient if payment_method == "hand_cash" else None
        if photo:
            existing_registration.photo = photo
        existing_registration.status = 'Pending'
        existing_registration.save()
        messages.success(request, "Your registration has been re-submitted successfully and is pending approval.")
    else:
        # Create new registration (starts as Pending)
        registration = EventRegistration.objects.create(
            user=request.user,
            event=event,
            name=name,
            student_id=student_id,
            email=email,
            phone=phone,
            payment_method=payment_method,
            transaction_id=transaction_id if payment_method == "bkash" else None,
            hand_cash_recipient=hand_cash_recipient if payment_method == "hand_cash" else None,
            photo=photo,
            status='Pending'
        )
        messages.success(request, "Your registration was submitted successfully and is pending approval.")
        
    return redirect("event_detail", event_id=event_id)


@login_required
def download_ticket_pdf(request, registration_id):
    """Generates and downloads the registration ticket stub as a PDF."""
    registration = get_object_or_404(EventRegistration, id=registration_id)
    
    # Restrict access to ticket owner or admin
    if registration.user != request.user and not is_admin(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("You do not have permission to download this ticket.")
        
    try:
        pdf_content = generate_ticket_pdf(registration)
        response = HttpResponse(pdf_content, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="Ticket_{registration.id}.pdf"'
        return response
    except Exception as e:
        messages.error(request, f"Error generating PDF ticket: {e}")
        return redirect("event_detail", event_id=registration.event.id)


@login_required
@user_passes_test(is_admin)
def approve_registration(request, registration_id):
    """Approves a pending registration and generates its ticket."""
    registration = get_object_or_404(EventRegistration, id=registration_id)
    if registration.status != 'Approved':
        registration.status = 'Approved'
        registration.save() # This generates serial_no and qr_code
        
        # Refresh to get QR code name, etc.
        registration.refresh_from_db()
        
        # Send confirmation email with attached PDF ticket
        subject = f"🎟 Your Event Ticket is Ready for {registration.event.title}"
        body = (
            f"Hello {registration.name or registration.user.get_full_name() or registration.user.username},\n\n"
            f"Your registration for '{registration.event.title}' has been APPROVED!\n"
            f"Your digital ticket has been generated.\n\n"
            f"Please find your digital ticket attached as a PDF to this email. Present this ticket at the event entrance.\n\n"
            f"Thanks,\n"
            f"BAIUST Robotics Society"
        )
        
        from django.core.mail import EmailMessage
        recipient_email = registration.email if registration.email else registration.user.email
        email_msg = EmailMessage(
            subject,
            body,
            settings.DEFAULT_FROM_EMAIL,
            [recipient_email],
        )
        
        try:
            pdf_content = generate_ticket_pdf(registration)
            email_msg.attach(f"Ticket_{registration.id}.pdf", pdf_content, "application/pdf")
        except Exception as e:
            print(f"Error generating PDF ticket attachment: {e}")
            
        email_msg.send(fail_silently=True)
        messages.success(request, f"Registration for {registration.name or registration.user.username} has been approved and ticket has been emailed.")
    else:
        messages.info(request, f"Registration is already approved.")
        
    return redirect("admin_dashboard")


@login_required
@user_passes_test(is_admin)
def reject_registration(request, registration_id):
    """Rejects a registration."""
    registration = get_object_or_404(EventRegistration, id=registration_id)
    if registration.status != 'Rejected':
        registration.status = 'Rejected'
        registration.save() # This clears serial_no and qr_code
        messages.warning(request, f"Registration for {registration.name or registration.user.username} has been rejected.")
    else:
        messages.info(request, f"Registration is already rejected.")
        
    return redirect("admin_dashboard")

